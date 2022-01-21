import sys

from openpyxl import load_workbook


recordsDict = {}
newRecordsDict = {}
addedRecords = []
modifiedRecords = []


# 读取xlxs内容
def loadXlsx(recordsDict, filename='adhub.xlsx'):
    wb = load_workbook(filename)
    ws = wb['Sheet']

    for i in range(2, ws.max_row + 1):
        # path
        curPath = ws.cell(row=i, column=4).value if ws.cell(
            row=i, column=4).value else ''
        key = ws.cell(row=i, column=1).value if ws.cell(
            row=i, column=1).value else ''
        cn = ws.cell(row=i, column=2).value if ws.cell(
            row=i, column=2).value else key
        en = ws.cell(row=i, column=3).value if ws.cell(
            row=i, column=3).value else key

        pathKey = "{curPath}@@@@@{key}".format(curPath=curPath, key=key)
        enCN = "{en}-{cn}".format(en=en, cn=cn)
        recordsDict[pathKey] = enCN

    wb.close()

# 比较新旧文件，得到变化的记录
def compareXlsxFiles():
    for (key, value) in newRecordsDict.items():
        if key in recordsDict.keys():
            if recordsDict[key] == value:
                continue
            else:
                modifiedRecords.append(key)
        else:
            addedRecords.append(key)

# 读取xlxs内容
def updateXlsx(filename='adhub.xlsx'):
    wb = load_workbook(filename)
    ws = wb['Sheet']

    ws['E1'] = '修改类型'

    for i in range(2, ws.max_row + 1):
        # path
        curPath = ws.cell(row=i, column=4).value if ws.cell(
            row=i, column=4).value else ''
        key = ws.cell(row=i, column=1).value if ws.cell(
            row=i, column=1).value else ''

        pathKey = "{curPath}@@@@@{key}".format(curPath=curPath, key=key)

        if pathKey in addedRecords:
            ws.cell(row=i, column=5).value = 'A'
        elif pathKey in modifiedRecords:
            ws.cell(row=i, column=5).value = 'M'

    wb.save(filename)
    wb.close()

def excetue():
    if len(sys.argv) != 3:
        print('参数个数不对')
        return

    xlxsFile = str(sys.argv[1])

    newXlxsFile = str(sys.argv[2])

    loadXlsx(recordsDict, xlxsFile)
    loadXlsx(newRecordsDict, newXlxsFile)
    compareXlsxFiles()
    updateXlsx(newXlxsFile)

    print('--finished--')


if __name__ == '__main__':
    excetue()
