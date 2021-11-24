import os
import sys

from openpyxl import load_workbook


recordsDict = {}


class Record(object):
    # 初始化中给对象属性赋值
    def __init__(self, key, value):
        self.key = key
        self.value = value


# 读取xlxs内容


def loadXlsx(filename='adhub.xlsx'):
    wb = load_workbook(filename)
    ws = wb['Sheet']

    for i in range(2, ws.max_row + 1):
        # path
        curPath = ws.cell(row=i, column=4).value
        key = ws.cell(row=i, column=1).value
        cn = ws.cell(row=i, column=2).value
        en = ws.cell(row=i, column=3).value

        if not curPath in recordsDict:
            recordsDict[curPath] = []

        recordsDict[curPath].append(Record(key, en))

        curZhPath = curPath.replace('en-US', 'zh-CN')
        if not curZhPath in recordsDict:
            recordsDict[curZhPath] = []

        recordsDict[curZhPath].append(Record(key, cn))

    wb.close()


def updateLocaleFiles():
    for (key, value) in recordsDict.items():
        createFile(key, value)

# 创建ts/js文件，并写入locales内容
def createFile(fileName, fileRecords):

    if not os.path.exists(os.path.dirname(fileName)):
        os.makedirs(os.path.dirname(fileName))

    if not os.path.exists(fileName):
        os.system(r"touch {}".format(fileName))  # 调用系统命令行来创建文件

    # 读写打开一个UTF-8编码格式文件，如果文件不存在则创建
    file = open(fileName, mode='w+', encoding='UTF-8')
    # open()打开一个文件，返回一个文件对象
    file.write('export default {\n')  # 写入文件

    for r in fileRecords:
        file.write("  '{key}': '{value}',\n".format(
            key=r.key, value=r.value))  # 写入文件

    file.write('}')  # 写入文件
    file.close()  # 关闭文件

def excetue():
    if len(sys.argv) != 2:
        print('参数个数不对')
        return
    
    xlxsFile = str(sys.argv[1])

    loadXlsx(xlxsFile)
    updateLocaleFiles()
    
    print('--finished--')

if __name__ == '__main__':
    excetue()
